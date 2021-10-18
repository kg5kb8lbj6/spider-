import time
from selenium import webdriver
import re
import requests
from pyquery import PyQuery
from openpyxl import workbook

url = 'https://club.xywy.com/list_331_all_1.htm'
headers = {'User-Agent':
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
            }
res = requests.get(url,headers = headers)
#print(res.status_code) #如果打印200则是请求成功
res.encoding = 'gbk'
html_data = res.text
excel = workbook.Workbook()#创建一个excel文件
sheet = excel.active   #获取当前wb的第一个worksheet，默认的索引值是0，它是可以改变的
data_list = [] #问题                  http://ask.seeys.com/
url_list = [] #问题的链接
sheet.cell(1, 1).value = '问题'
sheet.cell(1,5).value = '链接'

doc =  PyQuery((''.join(html_data)))
keyword = []
keyword_link = []
keyword1 = []
keyword_link1 = []
for item in doc.items('td.pl20.w340 a'):
    keyword.append(item.text())
    keyword_link.append(item.attr.href)

for i in range(1,len(keyword),2):
    keyword1.append(keyword[i])
#print(keyword1)
for i in range(1,len(keyword_link),2):
    keyword_link1.append(keyword_link[i])
#print(keyword_link1)

for i in range(0,len(keyword1)):
    sheet.cell(i + 2,1).value = keyword1[i]
    sheet.cell(i + 2,5).value = keyword_link1[i]

#第二页存数据
r = 2
for b in range(2,201):
    url1 = url = 'https://club.xywy.com/list_331_all_'+str(b)+'.htm'
    print(url1)
    res1 = requests.get(url1,headers = headers)
    res1.encoding = 'gbk'
    html_data1 = res1.text
    doc =  PyQuery((''.join(html_data1)))
    keyword = []
    keyword_link = []
    keyword1 = []
    keyword_link1 = []

    for item in doc.items('td.pl20.w340 a'):
        keyword.append(item.text())
        keyword_link.append(item.attr.href)

    for a in range(1,len(keyword),2):
        keyword1.append(keyword[a])
    #print(keyword1)
    #print(len(keyword1))
    for d in range(1,len(keyword_link),2):
        keyword_link1.append(keyword_link[d])
    #print(keyword_link1)
    #print(len(keyword_link1))

    for e in range(0,len(keyword1)):
        sheet.cell(20 * (b-1) + 2 + e,1).value = keyword1[e]
        sheet.cell(20 * (b-1) + 2 + e,5).value = keyword_link1[e]

    print('正在爬取第'+str(r)+'页的链接......')
    r += 1





excel.save(r'D:\xywyw.xlsx')