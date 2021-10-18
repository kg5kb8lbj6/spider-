import re

import pandas as pd
import openpyxl
import requests
from pyquery import PyQuery
import time
import random
headers = {'User-Agent':
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
            }
main_book = openpyxl.load_workbook(r'D:\xywyw3.xlsx')
main_sheet = main_book.active
main_sheet.cell(1,11).value = '详细询问'
main_sheet.cell(1,21).value = '医生建议'

r = 0
for i in range(3500,4001):
    res = requests.get(main_sheet.cell(i, 5).value, headers=headers)
    res.encoding = 'gbk'
    html_data = res.text
    doc = PyQuery((''.join(html_data)))
    keyword = []
    keyword1 = []
    for item in doc.items('div.graydeep.User_quecol.pt10.mt10'):
        keyword.append(item.text())
    print(keyword)
    main_sheet.cell(3500 + r,11).value = ''.join(keyword)


    for item in doc.items('div.pt15.f14.graydeep.pl20.pr20.deepblue'):
        keyword1.append(item.text())
    print(keyword1)
    main_sheet.cell(3500 + r,21).value = ''.join(keyword1)
    r += 1
    print(i)

time.sleep(random.randint(1,4))
main_book.save(r'D:\xywyw2.xlsx')